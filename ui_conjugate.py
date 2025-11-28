from openpyxl import load_workbook
from stem_utils import apply_stem_shift


def ui_conjugate_presente(verb_es):
    verb = UISpanishVerb(verb_es)
    persons = get_persons(verb)
    stem = verb.stemPresente
    endings = get_endings(verb)

    forms = [f"{person} {s}{ending}" for person, s, ending in zip(persons, stem, endings)]
    return "\n".join(forms)

def get_persons(verb):
    if verb.is_reflexive:
        return ["yo me", "tú te", "él/ella se", "nosotros nos", "vosotros os", "ellos/ellas se"]
    else:
        return ["yo", "tú", "él/ella", "nosotros", "vosotros", "ellos/ellas"]

class UISpanishVerb:

    def __init__(self, infinitive):
        self.infinitive = infinitive.strip().lower()
        self.is_reflexive = self.infinitive.endswith("se")
        self.group = self._get_group()
        self.irregular = self._get_irregular_tags()
        self.stemPresente = self._get_stem_presente()

    def is_reflexive_verb(self):
        return self.is_reflexive

    def _get_stem_presente(self):
        infinitive = self.infinitive
        if self.is_reflexive:
            infinitive = infinitive[:-2]
        if infinitive == "ir":
            return ["v"] * 6
        elif infinitive == "ser":
            return ["s", "er", "e", "s", "s", "s"]
        elif infinitive == "caber":
            return ["quep"] + [infinitive[:-2]] * 5
        elif infinitive == "saber":
            return ["s"] + [infinitive[:-2]] * 5
        elif "ig" in self.irregular:
            stem = infinitive[:-2]
            if "y" in self.irregular:
                stem_list = [stem + "ig", stem + "y", stem + "y", stem, stem, stem + "y"]
                return stem_list
            else:
                return [stem + "ig"] + [stem] * 5
        elif "zco" in self.irregular:
            stem = infinitive[:-2]
            return [stem[:-1] + "zc"] + [stem] * 5
        elif "ngo" in self.irregular:
            stem = infinitive[:-2]
            yo_form = stem + "g"
            if "eie" in self.irregular:
                stem_list = apply_stem_shift(stem, "e", "ie", [0, 1, 2, 5])
                stem_list[0] = yo_form
                return stem_list
            else: return [yo_form] + [stem] * 5
        elif "eie" in self.irregular:
            stem = infinitive[:-2]
            stem_list = apply_stem_shift(stem, "e", "ie", [0, 1, 2, 5])
            return stem_list
        elif "oue" in self.irregular:
            stem = infinitive[:-2]
            stem_list = apply_stem_shift(stem, "o", "ue", [0, 1, 2, 5])
            if "zo" in self.irregular:
                stem_list[0] = stem_list[0][:-1] + "z"
            return stem_list
        elif "j" in self.irregular:
            stem = infinitive[:-2]
            yo_form = infinitive[:-3] + "j"
            if "ei" in self.irregular:
                stem_list = apply_stem_shift(stem, "e", "i", [0, 1, 2, 5])
                stem_list[0] = stem_list[0][:-1] + "j"
                return stem_list
            else: return [yo_form] + [stem] * 5
        elif "ei" in self.irregular:
            stem = infinitive[:-2]
            stem_list = apply_stem_shift(stem, "e", "i", [0, 1, 2, 5])
            if "g" in self.irregular: stem_list[0] = stem_list[0][:-1] + "g"
            elif "um" in self.irregular: stem_list[0] = stem_list[0][:-1]
            return stem_list
        elif "g2" in self.irregular:
            stem = infinitive[:-2]
            return [stem[:-1] + "g"] + [stem] * 5
        elif "g" in self.irregular:
            stem = infinitive[:-2]
            return [stem + "g"] + [stem] * 5
        elif "zo" in self.irregular:
            stem = infinitive[:-2]
            return [stem[:-1] + "z"] + [stem] * 5
        elif "y" in self.irregular:
            stem = infinitive[:-2] + "y"
            stem_list = [stem, stem, stem, stem[:-1], stem[:-1], stem]
            return stem_list
        elif "ií" in self.irregular:
            stem = infinitive[:-2]
            stem_list = apply_stem_shift(stem, "i", "í", [0, 1, 2, 5])
            return stem_list
        elif "eí" in self.irregular:
            stem = infinitive[:-2]
            stem_list = apply_stem_shift(stem, "e", "í", [0, 1, 2, 5])
            return stem_list
        elif "uue" in self.irregular:
            stem = infinitive[:-2]
            stem_list = apply_stem_shift(stem, "u", "ue", [0, 1, 2, 5])
            return stem_list
        else:
            stem = infinitive[:-2]
            return [stem] * 6

    def _get_group(self):
        infinitive = self.infinitive

        if infinitive in ["ir", "irse"]:
            return "ar"
        elif infinitive.endswith(("ar", "arse")):
            return "ar"
        elif infinitive.endswith(("er", "erse")):
            return "er"
        else:
            return "ir"

    def _get_irregular_tags(self):
        verb_es = self.infinitive.strip().lower()
        wb = load_workbook("verbos.xlsx")
        ws = wb["verbos"]

        for row in ws.iter_rows(min_row=3):
            cell_value = str(row[1].value).strip().lower()
            if cell_value == verb_es:
                cell_c = row[2]
                if cell_c:
                    tags = str(cell_c.value).strip().lower().split(",")
                    return [tag.strip() for tag in tags]
                return []
        return []

def get_endings(verb):
    if verb.group == "ar":
        if "oy" in verb.irregular:
            if verb.stemPresente in ["v", "d"]:
                return ["oy", "as", "a", "amos", "ais", "an"]
            else:
                return ["oy", "as", "a", "amos", "áis", "an"]
        else:
            return ["o", "as", "a", "amos", "áis", "an"]
    elif verb.group == "er":
        if "oy" in verb.irregular:
            return ["oy", "es", "s", "omos", "ois", "on"]
        elif "e" in verb.irregular:
            return ["eo", "es", "e", "emos", "eis", "en"]
        elif verb.infinitive=="saber":
            return ["é", "es", "e", "emos", "eis", "en"]
        else:
            return ["o", "es", "e", "emos", "éis", "en"]
    else:
        if "eí" in verb.irregular:
            return ["o", "es", "e", "ímos", "ís", "en"]
        return ["o", "es", "e", "imos", "ís", "en"]