from openpyxl import load_workbook


def conjugate_presente(verb_es):
    verb = SpanishVerb(verb_es)
    persons = get_persons(verb)
    stem = verb.stemPresente
    def get_endings(verb):
        if verb.group == "ar":
            if "oy" in verb.irregular:
                if verb.stemPresente in ["v", "d"]:
                    return ["oy", "as", "a", "amos", "ais", "an"]
                else:
                    return ["oy", "as", "a", "amos", "áis", "an"]
            else:
                return ["o", "as", "a", "amos", "áis", "an"]
        elif verb.group == "er":
            if "oy" in verb.irregular:
                return ["oy", "es", "s", "omos", "ois", "on"]
            elif "e" in verb.irregular:
                return ["eo", "es", "e", "emos", "eis", "en"]
            else:
                return ["o", "es", "e", "emos", "éis", "en"]
        elif verb.group == "ir":
            if "eí" in verb.irregular:
                return ["o", "es", "e", "ímos", "ís", "en"]
            return ["o", "es", "e", "imos", "ís", "en"]
    endings = get_endings(verb)

    forms = [f"{person} {s}{ending}" for person, s, ending in zip(persons, stem, endings)]
    return "\n".join(forms)

def get_persons(verb):
    if verb.is_reflexive:
        return ["yo me", "tú te", "él/ella se", "nosotros nos", "vosotros os", "ellos/ellas se"]
    else:
        return ["yo", "tú", "él/ella", "nosotros", "vosotros", "ellos/ellas"]

class SpanishVerb:

    def __init__(self, infinitive):
        self.infinitive = infinitive.strip().lower()
        self.is_reflexive = self.infinitive.endswith("se")
        self.group = self._get_group()
        self.irregular = self._get_irregular_tags()
        self.stemPresente = self._get_stem_presente()

    def is_reflexive_verb(self):
        return self.is_reflexive

    def _get_stem_presente(self):
        infinitive = self.infinitive
        if self.is_reflexive:
            infinitive = infinitive[:-2]
        if infinitive == "ir":
            return ["v"] * 6
        elif infinitive == "ser":
            return ["s", "er", "e", "s", "s", "s"]
        elif infinitive == "oler":
            return ["huel", "huel", "huel", "ol", "ol", "huel"]
        if infinitive == "caber":
            return ["quep"] + [infinitive[:-2]] * 5
        elif "ig" in self.irregular:
            stem = infinitive[:-2]
            if "y" in self.irregular:
                stem_list = [stem + "ig", stem + "y", stem + "y", stem, stem, stem + "y"]
                return stem_list
            else:
                return [stem + "ig"] + [stem] * 5
        elif "zco" in self.irregular:
            stem = infinitive[:-2]
            return [stem[:-1] + "zc"] + [stem] * 5
        elif "ngo" in self.irregular:
            stem = infinitive[:-2]
            yo_form = stem + "g"
            if "eie" in self.irregular:
                def apply_ie_shift(s):
                    index = s.rfind("e")
                    return s[:index] + "ie" + s[index + 1:] if index != -1 else s
                stem_list = [yo_form, apply_ie_shift(stem), apply_ie_shift(stem), stem, stem, apply_ie_shift(stem)]
                return stem_list
            else: return [yo_form] + [stem] * 5
        elif "eie" in self.irregular:
            stem = infinitive[:-2]
            def apply_ie_shift(s):
                index = s.rfind("e")
                return s[:index] + "ie" + s[index + 1:] if index != -1 else s
            stem_list = [apply_ie_shift(stem), apply_ie_shift(stem), apply_ie_shift(stem), stem, stem, apply_ie_shift(stem)]
            return stem_list
        elif "oue" in self.irregular:
            yo_form =stem = infinitive[:-2]
            if "zo" in self.irregular: yo_form = infinitive[:-3] + "z"
            def apply_ue_shift(s):
                index = s.rfind("o")
                return s[:index] + "ue" + s[index + 1:] if index != -1 else s
            stem_list = [apply_ue_shift(yo_form), apply_ue_shift(stem), apply_ue_shift(stem), stem, stem, apply_ue_shift(stem)]
            return stem_list
        elif "j" in self.irregular:
            stem = infinitive[:-2]
            yo_form = infinitive[:-3] + "j"
            if "ei" in self.irregular:
                def apply_i_shift(s):
                    index = s.rfind("e")
                    return s[:index] + "i" + s[index + 1:] if index != -1 else s
                stem_list = [apply_i_shift(yo_form), apply_i_shift(stem), apply_i_shift(stem), stem, stem, apply_i_shift(stem)]
                return stem_list
            else: return [yo_form] + [stem] * 5
        elif "ei" in self.irregular:
            stem = infinitive[:-2]
            yo_form = infinitive[:-2]
            if "g" in self.irregular: yo_form = stem[:-1] + "g"
            elif "um" in self.irregular: yo_form = stem[:-1]
            def apply_ie_shift(s):
                index = s.rfind("e")
                return s[:index] + "i" + s[index + 1:] if index != -1 else s
            stem_list = [apply_ie_shift(yo_form), apply_ie_shift(stem), apply_ie_shift(stem), stem, stem, apply_ie_shift(stem)]
            return stem_list
        elif "g2" in self.irregular:
            stem = infinitive[:-2]
            return [stem[:-1] + "g"] + [stem] * 5
        elif "g" in self.irregular:
            stem = infinitive[:-2]
            return [stem + "g"] + [stem] * 5
        elif "zo" in self.irregular:
            stem = infinitive[:-2]
            return [stem[:-1] + "z"] + [stem] * 5
        elif "y" in self.irregular:
            stem = infinitive[:-2] + "y"
            stem_list = [stem, stem, stem, stem[:-1], stem[:-1], stem]
            return stem_list
        elif "ií" in self.irregular:
            stem = infinitive[:-2]
            def apply_ii_shift(s):
                index = s.rfind("i")
                return s[:index] + "í" + s[index + 1:] if index != -1 else s
            stem_list = [apply_ii_shift(stem), apply_ii_shift(stem), apply_ii_shift(stem), stem, stem, apply_ii_shift(stem)]
            return stem_list
        elif "eí" in self.irregular:
            stem = infinitive[:-2]
            def apply_ei_shift(s):
                index = s.rfind("e")
                return s[:index] + "í" + s[index + 1:] if index != -1 else s
            stem_list = [apply_ei_shift(stem), apply_ei_shift(stem), apply_ei_shift(stem), stem, stem, apply_ei_shift(stem)]
            return stem_list
        else:
            stem = infinitive[:-2]
            return [stem] * 6


    def _get_group(self):
        infinitive = self.infinitive

        if infinitive in ["ir", "irse"]:
            return "ar"
        elif infinitive.endswith(("ar", "arse")):
            return "ar"
        elif infinitive.endswith(("er", "erse")):
            return "er"
        else:
            return "ir"

    def _get_irregular_tags(self):
        verb_es = self.infinitive.strip().lower()
        wb = load_workbook("verbos.xlsx")
        ws = wb["verbos"]

        for row in ws.iter_rows(min_row=3):
            cell_value = str(row[1].value).strip().lower()
            if cell_value == verb_es:
                cell_c = row[2]
                if cell_c:
                    tags = str(cell_c.value).strip().lower().split(",")
                    return [tag.strip() for tag in tags]
                return []
        return []